import datetime
import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl

class Usuario:
    def __init__(self, nome, cpf, setor, localizacao, assinatura, celular, computador=None):
        self.nome = nome 
        self.cpf = cpf
        self.setor = setor
        self.localizacao = localizacao
        self.assinatura = assinatura 
        self.celular = celular 
        self.computador = computador
       
class Computador:
    def __init__(self, numero, serial, nome_modelo, obs, usuario=None, data_alocacao=None):
        self.numero = numero
        self.serial = serial
        self.nome_modelo = nome_modelo
        self.obs = obs
        self.usuario = usuario
        self.data_alocacao = data_alocacao
        self.registros = [] 

class Registro:
    def __init__(self, numero, data_alocacao, data_devolucao, ultimo_usuario_alocou):
        self.numero = numero
        self.data_alocacao = data_alocacao
        self.data_devolucao = data_devolucao
        self.ultimo_usuario_alocou = ultimo_usuario_alocou
      
class GerenciadorComputadores:
    def __init__(self):
        self.computadores = []

    def adicionar_computador(self, computador):
        self.computadores.append(computador)

    def computadores_disponiveis(self):
        return [comp for comp in self.computadores if comp.usuario is None]

class GerenciadorAlocados:
    def __init__(self):
        self.alocados = []
        self.registros = [] 

    def adicionar_alocado(self, computador):
        self.alocados.append(computador)

    def remover_alocado(self, computador):
        self.alocados.remove(computador)
        registro_desalocado = Registro(computador.numero, computador.data_alocacao, datetime.datetime.now(), computador.usuario.nome if computador.usuario else None)
        self.registros.append(registro_desalocado)

    def computadores_alocados(self):
        return self.alocados

# Funções para salvar e carregar dados em um arquivo Excel
def salvar_usuarios_computadores_excel(filename, usuarios, computadores):
    workbook = openpyxl.Workbook()
    sheet_usuarios = workbook.create_sheet(title="Usuários")
    sheet_usuarios.append(["Nome", "CPF", "Setor", "Localizacao", "Assinatura", "Celular"])

    for usuario in usuarios:
        sheet_usuarios.append([usuario.nome, usuario.cpf, usuario.setor, usuario.localizacao, usuario.assinatura, usuario.celular])

    sheet_computadores = workbook.create_sheet(title="Computadores")
    sheet_computadores.append(["Número", "Serial", "Nome/Modelo", "Observações"])

    for computador in computadores:
        sheet_computadores.append([computador.numero, computador.serial, computador.nome_modelo, computador.obs])

    workbook.save(filename)

def salvar_computador_alocado_excel(filename, computadores_alocados):
    workbook = openpyxl.load_workbook(filename)
    # Se a folha "Alocados" já existir, remove-a antes de adicionar novamente
    if "Alocados" in workbook.sheetnames:
        del workbook["Alocados"]
    sheet_alocados = workbook.create_sheet(title="Alocados")

    sheet_alocados.append(["Número", "Serial", "Nome/Modelo", "Usuário", "Observação", "Assinatura", "Celular"])

    for computador in computadores_alocados:
        if computador.usuario:
            sheet_alocados.append([computador.numero, computador.serial, computador.nome_modelo,
                                   computador.usuario.nome, computador.obs, computador.usuario.assinatura, computador.usuario.celular])

    workbook.save(filename)

# Atualizar a função desalocar_computador para lidar com o novo campo de celular e assinatura

def carregar_usuarios_computadores_excel(filename):
    usuarios = []
    computadores = []
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet_usuarios = workbook["Usuários"]

        for row in sheet_usuarios.iter_rows(min_row=2, values_only=True):
            nome, cpf, setor, localizacao, assinatura, celular = row
            usuarios.append(Usuario(nome, cpf, setor, localizacao, assinatura, celular))

        sheet_computadores = workbook["Computadores"]

        for row in sheet_computadores.iter_rows(min_row=2, values_only=True):
            numero, serial, nome_modelo, obs = row
            computadores.append(Computador(numero, serial, nome_modelo, obs))
    except FileNotFoundError:
        pass

    return usuarios, computadores
def carregar_computadores_alocados_excel(filename, lista_usuarios):
    computadores_alocados = []
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet_alocados = workbook["Alocados"]

        for row in sheet_alocados.iter_rows(min_row=2, values_only=True):
            if len(row) == 7:
                numero, serial, nome_modelo, usuario_nome, observacao, assinatura, celular = row
                usuario = next((u for u in lista_usuarios if u.nome == usuario_nome), None)
                if usuario:
                    observacao = ""
                    computador = Computador(numero, serial, nome_modelo, observacao, usuario)
                    computadores_alocados.append(computador)
            else:
                print(f"Aviso: Linha ignorada no arquivo {filename}. Não possui informações suficientes.")

    except FileNotFoundError:
        pass
    except Exception as e:
        print(f"Erro ao carregar dados do arquivo {filename}: {e}")

    return computadores_alocados



def carregar_registros_excel(filename):
    registros = []
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet_registros = workbook["Registros"]

        for row in sheet_registros.iter_rows(min_row=2, values_only=True):
            numero, data_alocacao, data_devolucao, ultimo_usuario_alocou = row
            registro = Registro(numero, data_alocacao, data_devolucao, ultimo_usuario_alocou)
            registros.append(registro)
    except KeyError:
      
        workbook.save(filename)  # Salva as alterações no arquivo
    except FileNotFoundError:
        pass

    return registros

# Caminho do arquivo Excel
excel_file = "C:/Users/user/Desktop/Patrimonio.xlsx"

# Carregar dados do Excel ou criar arquivo vazio se não existir
lista_usuarios, lista_computadores = carregar_usuarios_computadores_excel(excel_file)

# Carregar dados dos computadores alocados
lista_computadores_alocados = carregar_computadores_alocados_excel(excel_file, lista_usuarios)

# Instanciar GerenciadorComputadores e GerenciadorAlocados no início
gerenciador_computadores = GerenciadorComputadores()
gerenciador_alocados = GerenciadorAlocados()

# Adicionar os computadores alocados carregados ao gerenciador_alocados
for comp in lista_computadores_alocados:
    gerenciador_alocados.adicionar_alocado(comp)

# Adicionar os computadores carregados ao gerenciador_computadores
for comp in lista_computadores:
    # Verificar se o computador está alocado para um usuário
    if comp.usuario:
        gerenciador_alocados.adicionar_alocado(comp)
    else:
        gerenciador_computadores.adicionar_computador(comp)

# Funções para a interface gráfica
def cadastrar_usuario():
    nome = entry_nome.get()
    cpf = entry_cpf.get()

    # Verificar se o CPF já foi cadastrado
    if any(u.cpf == cpf for u in lista_usuarios):
        messagebox.showwarning("Aviso", "Este CPF já foi cadastrado.")
        return

    setor = setor_var.get()
    localizacao = localizacao_var.get()
    assinatura = assinatura_menu.cget("text")  # Obter o texto do item selecionado no menu
    celular = celular_var.get()

    usuario = Usuario(nome, cpf, setor, localizacao, assinatura, celular)
    lista_usuarios.append(usuario)  # Adiciona o usuário à lista de usuários

    # Limpar os campos após o cadastro
    entry_nome.delete(0, tk.END)
    entry_cpf.delete(0, tk.END)

    messagebox.showinfo("Sucesso", "Usuário cadastrado com sucesso!")
    atualizar_comboboxes_alocacao()

    # Salva usuários e computadores no arquivo Excel
    salvar_usuarios_computadores_excel(excel_file, lista_usuarios, gerenciador_computadores.computadores)
    salvar_computador_alocado_excel(excel_file, gerenciador_alocados.computadores_alocados())


def cadastrar_computador():
    numero = entry_numero.get()
    serial = entry_serial.get()
    nome_modelo = entry_nome_modelo.get()
    obs = entry_obs.get()

    computador = Computador(numero, serial, nome_modelo, obs)
    gerenciador_computadores.adicionar_computador(computador)

    # Limpar os campos após o cadastro
    entry_numero.delete(0, tk.END)
    entry_serial.delete(0, tk.END)
    entry_nome_modelo.delete(0, tk.END)
    entry_obs.delete(0, tk.END)

    messagebox.showinfo("Sucesso", "Computador cadastrado com sucesso!")
    atualizar_comboboxes_alocacao()

    # Salva usuários e computadores no arquivo Excel
    salvar_usuarios_computadores_excel(excel_file, lista_usuarios, gerenciador_computadores.computadores)
    salvar_computador_alocado_excel(excel_file, gerenciador_alocados.computadores_alocados())

def mostrar_computadores_disponiveis():
    disponiveis = gerenciador_computadores.computadores_disponiveis()
    texto = "Computadores Disponíveis:\n"
    for comp in disponiveis:
        texto += f"Número: {comp.numero}, Série: {comp.serial}, Nome/Modelo: {comp.nome_modelo}\n"
    messagebox.showinfo("Computadores Disponíveis", texto)
def alocar_computador():
    # Obter o usuário selecionado
    usuario_selecionado = lista_usuarios_combobox.get()

    if not usuario_selecionado:
        messagebox.showwarning("Aviso", "Selecione um usuário para alocar o computador.")
        return

    # Obter o computador disponível selecionado
    comp_disponivel_selecionado = lista_computadores_combobox.get()

    if not comp_disponivel_selecionado:
        messagebox.showwarning("Aviso", "Selecione um computador disponível para alocar ao usuário.")
        return

    # Encontrar o objeto Usuario correspondente ao nome selecionado
    usuario = next((u for u in lista_usuarios if u.nome == usuario_selecionado), None)

    # Encontrar o objeto Computador correspondente ao número selecionado
    comp_disponivel = next((comp for comp in gerenciador_computadores.computadores_disponiveis() if comp.serial == comp_disponivel_selecionado), None)

    if not usuario or not comp_disponivel:
        messagebox.showerror("Erro", "Erro ao alocar o computador.")
        return

    # Alocar o computador ao usuário
    comp_disponivel.usuario = usuario
    usuario.computador = comp_disponivel

    # Registrar a data de alocação
    comp_disponivel.data_alocacao = datetime.datetime.now()

    # Adicionar os dados de assinatura e celular ao usuário
    assinatura = assinatura_var.get()
    celular = celular_var.get()
    usuario.assinatura = assinatura
    usuario.celular = celular

    # Remover o computador da lista de computadores disponíveis
    gerenciador_computadores.computadores.remove(comp_disponivel)

    # Adicionar o computador à lista de computadores alocados
    gerenciador_alocados.adicionar_alocado(comp_disponivel)

    # Atualizar os comboboxes

    # Salvar as alterações no Excel
    salvar_usuarios_computadores_excel(excel_file, lista_usuarios, gerenciador_computadores.computadores)
    salvar_computador_alocado_excel(excel_file, gerenciador_alocados.computadores_alocados())
    atualizar_comboboxes_alocacao()

    messagebox.showinfo("Sucesso", f"Computador alocado para {usuario.nome} em {comp_disponivel.data_alocacao}.")

# Corrigindo a função desalocar_computador para registrar corretamente a desalocação

def desalocar_computador():
    # Obter o computador alocado selecionado
    comp_alocado_selecionado = lista_computadores_alocados_combobox.get()

    if not comp_alocado_selecionado:
        messagebox.showwarning("Aviso", "Selecione um computador alocado para desalocar.")
        return

    # Encontrar o objeto Computador correspondente ao número selecionado
    comp_alocado = next((comp for comp in gerenciador_alocados.computadores_alocados() if comp.serial == comp_alocado_selecionado), None)

    if not comp_alocado:
        messagebox.showerror("Erro", "Erro ao desalocar o computador.")
        return

    # Verificar se o usuário associado ao computador alocado não é None
    if comp_alocado.usuario is not None:
        # Atualizar informações do computador
        comp_alocado.usuario.computador = None
        comp_alocado.data_devolucao = datetime.datetime.now()
        comp_alocado.ultimo_usuario_alocou = comp_alocado.usuario.nome if comp_alocado.usuario else None

        # Remover o computador da lista de computadores alocados
        gerenciador_alocados.remover_alocado(comp_alocado)

        # Adicionar o computador à lista de computadores disponíveis
        gerenciador_computadores.adicionar_computador(comp_alocado)

        # Registrar a data de devolução
        comp_alocado.data_devolucao = datetime.datetime.now()

        salvar_usuarios_computadores_excel(excel_file, lista_usuarios, gerenciador_computadores.computadores)
        salvar_computador_alocado_excel(excel_file, gerenciador_alocados.computadores_alocados())

        # Salvar registro após desalocação
        registro_desalocado = Registro(comp_alocado.numero, comp_alocado.data_alocacao, comp_alocado.data_devolucao, comp_alocado.ultimo_usuario_alocou)
        gerenciador_alocados.registros.append(registro_desalocado)
        salvar_registros_excel(excel_file, gerenciador_alocados.registros)

        # Remover o computador desalocado dos comboboxes
        lista_computadores_alocados_combobox.set('')
        lista_computadores_alocados_combobox['values'] = [comp.serial for comp in gerenciador_alocados.computadores_alocados()]
        lista_computadores_combobox['values'] = [comp.serial for comp in gerenciador_computadores.computadores_disponiveis()]

        # Atualizar a contabilização de computadores disponíveis
        atualizar_quantidades()

        messagebox.showinfo("Sucesso", f"Computador desalocado.")
    else:
        messagebox.showerror("Erro", "Erro ao desalocar o computador: Usuário associado é None.")

    # Atualizar os comboboxes
    atualizar_comboboxes_alocacao()


def atualizar_quantidades():
    quant_disponiveis = len(gerenciador_computadores.computadores_disponiveis())
    quant_alocados = len(gerenciador_alocados.computadores_alocados())
    total_computadores = quant_disponiveis + quant_alocados
    label_quantidade_disponiveis.config(text=f"Computadores Disponíveis: {quant_disponiveis}")
    label_quantidade_alocados.config(text=f"Computadores Alocados: {quant_alocados}")
    label_total_computadores.config(text=f"Total de Computadores: {total_computadores}")




def atualizar_comboboxes_alocacao():
    # Limpar as listas dos comboboxes de alocação
    lista_usuarios_combobox['values'] = [u.nome for u in lista_usuarios]
    lista_computadores_combobox['values'] = [comp.serial for comp in gerenciador_computadores.computadores_disponiveis()]
    lista_computadores_alocados_combobox['values'] = [comp.serial for comp in gerenciador_alocados.computadores_alocados()]

    # Atualizar os controles de quantidade de computadores
    atualizar_quantidades()

def salvar_registros_excel(filename, registros):
    workbook = openpyxl.load_workbook(filename)
    sheet_registros = workbook.create_sheet(title="Registros")

    # Adicione cabeçalhos
    sheet_registros.append(["Número", "Data Alocação", "Data Devolução", "Último Usuário Alocou"])

    # Adicione registros
    for registro in registros:
        sheet_registros.append([registro.numero, registro.data_alocacao, registro.data_devolucao, registro.ultimo_usuario_alocou])

    workbook.save(filename)

# Interface gráfica
root = tk.Tk()
root.title("Cadastro de Usuários e Computadores")

notebook = ttk.Notebook(root)

tab_usuarios = ttk.Frame(notebook)
notebook.add(tab_usuarios, text="Usuários")

label_nome = tk.Label(tab_usuarios, text="Nome:")
label_nome.pack()
entry_nome = tk.Entry(tab_usuarios)
entry_nome.pack()

label_cpf = tk.Label(tab_usuarios, text="CPF:")
label_cpf.pack()
entry_cpf = tk.Entry(tab_usuarios)
entry_cpf.pack()

label_setor = tk.Label(tab_usuarios, text="Setor:")
label_setor.pack()
setores_usuario = ["Administração", "TI", "RH", "Comercial", "Financeiro", "Manunteção"]
setor_var = tk.StringVar(tab_usuarios)
setor_var.set(setores_usuario[0])
setor_menu = tk.OptionMenu(tab_usuarios, setor_var, *setores_usuario)
setor_menu.pack()

label_localizacao = tk.Label(tab_usuarios, text="Localizacao:")
label_localizacao.pack()
localizacao_usuario = ["Campinas", "São Paulo", "Home-office"]
localizacao_var = tk.StringVar(tab_usuarios)
localizacao_var.set(localizacao_usuario[0])
localizacao_menu = tk.OptionMenu(tab_usuarios, localizacao_var, *localizacao_usuario)
localizacao_menu.pack()

btn_cadastrar_usuario = tk.Button(tab_usuarios, text="Cadastrar Usuário", command=cadastrar_usuario)
btn_cadastrar_usuario.pack()

tab_computadores = ttk.Frame(notebook)
notebook.add(tab_computadores, text="Computadores")

label_numero = tk.Label(tab_computadores, text="Número do Computador:")
label_numero.pack()
entry_numero = tk.Entry(tab_computadores)
entry_numero.pack()

label_serial = tk.Label(tab_computadores, text="Número de Série:")
label_serial.pack()
entry_serial = tk.Entry(tab_computadores)
entry_serial.pack()

label_nome_modelo = tk.Label(tab_computadores, text="Nome/Modelo:")
label_nome_modelo.pack()
entry_nome_modelo = tk.Entry(tab_computadores)
entry_nome_modelo.pack()

label_obs = tk.Label(tab_computadores, text="Observações:")
label_obs.pack()
entry_obs = tk.Entry(tab_computadores)
entry_obs.pack()

btn_cadastrar_computador = tk.Button(tab_computadores, text="Cadastrar Computador", command=cadastrar_computador)
btn_cadastrar_computador.pack()

tab_alocacao = ttk.Frame(notebook)
notebook.add(tab_alocacao, text="Alocação")

label_usuarios_combobox = tk.Label(tab_alocacao, text="Selecione o Usuário:")
label_usuarios_combobox.pack()
lista_usuarios_combobox = ttk.Combobox(tab_alocacao, values=[], state="readonly")
lista_usuarios_combobox.pack()

label_computadores_combobox = tk.Label(tab_alocacao, text="Selecione o Computador Disponível:")
label_computadores_combobox.pack()
lista_computadores_combobox = ttk.Combobox(tab_alocacao, values=[], state="readonly")
lista_computadores_combobox.pack()

label_assinatura = tk.Label(tab_alocacao, text="Assinatura do Termo :")
label_assinatura.pack()
assinatura_usuario = ["Sim", "Não"]
assinatura_var = tk.StringVar(tab_alocacao)
assinatura_var.set(assinatura_usuario[0])
assinatura_menu = tk.OptionMenu(tab_alocacao, assinatura_var, *assinatura_usuario)
assinatura_menu.pack()

label_celular = tk.Label(tab_alocacao, text="Celular:")
label_celular.pack()
celular_usuario = ["Nem Um", "Celular", "Mouse", "Teclado", "Fone", "Suporte", "Teclado/Mouse", "Teclado/Mouse/Suporte/Fone",
                    "Teclado/Mouse/Suporte", "Celular/Teclado/Mouse/Suporte/", "Celular/Teclado/Mouse", "Celular/Mouse",
                    "Celular/Teclado"]
celular_var = tk.StringVar(tab_alocacao)
celular_var.set(celular_usuario[0])
celular_menu = tk.OptionMenu(tab_alocacao, celular_var, *celular_usuario)
celular_menu.pack()


btn_alocar_computador = tk.Button(tab_alocacao, text="Alocar Computador", command=alocar_computador)
btn_alocar_computador.pack()

label_quantidade_disponiveis = tk.Label(tab_alocacao, text="")
label_quantidade_disponiveis.pack()

label_quantidade_alocados = tk.Label(tab_alocacao, text="")
label_quantidade_alocados.pack()

label_total_computadores = tk.Label(tab_alocacao, text="")
label_total_computadores.pack()

tab_excluir = ttk.Frame(notebook)
notebook.add(tab_excluir, text="Excluir")

label_computadores_alocados_combobox = tk.Label(tab_excluir, text="Selecione o Computador Alocado:")
label_computadores_alocados_combobox.pack()
lista_computadores_alocados_combobox = ttk.Combobox(tab_excluir, values=[], state="readonly")
lista_computadores_alocados_combobox.pack()

btn_desalocar_computador = tk.Button(tab_excluir, text="Desalocar Computador", command=desalocar_computador)
btn_desalocar_computador.pack()

atualizar_comboboxes_alocacao()
atualizar_quantidades()

notebook.pack()

root.protocol("WM_DELETE_WINDOW", lambda: (root.destroy()))
root.mainloop() 
